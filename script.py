#!/usr/bin/python
# -*- coding: UTF-8 -*-

import os
import sys 
import re
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import ConfigParser

reload(sys)
sys.setdefaultencoding( "utf-8" )

class obj:
	def __init__(self):
		self.mbl_no = ""
		self.hbl_no = ""
		self.containers = []
		self.cargo_descriptions = []
		self.hts_codes = []
		self.marks = []
		self.mark_merge = False
		
class container:
	def __init__(self):
		self.container_number = ""
		self.seal_number = ""
		self.amount = 0
		self.type = ""
		self.gross_weight = 0
		self.volume = 0

def remove_BOM(config_path):
	content = open(config_path).read()
	content = re.sub(r"\xfe\xff","", content)
	content = re.sub(r"\xff\xfe","", content)
	content = re.sub(r"\xef\xbb\xbf","", content)
	open(config_path, 'w').write(content)		

def getMaxLength(column):
	len_list = []
	for row in column:
		length = 0
		if row.value != None:
			elem_split = list(str(row.value).decode("utf-8"))
			for c in elem_split:
				if ord(c) <= 256:
					length += 1
				else:
					length += 2
			len_list.append(length)
	return max(len_list)

def autoWidth(sh):
	columns = sh.columns
	i = 0
	fill = PatternFill("solid", fgColor="92D050")
	fontObj = Font(name='Maersk Text', bold=True, italic=False, size=10)
	for column in columns:
		sh.cell(1, i+1).font = fontObj  # 加粗表头
		sh.cell(1, i+1).alignment = Alignment(horizontal='center')
		sh.cell(1, i+1).fill = fill
		max_len = getMaxLength(column)
		letter = chr(i+65)
		if max_len <= 12:
			sh.column_dimensions[letter].width = max(12, max_len + 4)
		else:
			sh.column_dimensions[letter].width = max_len + 4
		i = i + 1
		
	
def inMerged(row,col,sh):
	for merged in sh.merged_cells:
		if (row >= merged[0] and row < merged[1] and col >= merged[2] and col < merged[3]):
			return True
	return False

def cell_real_value(row,col,sh):
	for merged in sh.merged_cells:
		if (row >= merged[0] and row < merged[1] and col >= merged[2] and col < merged[3]):
			return sh.cell_value(merged[0],merged[2])
	return sh.cell_value(row,col)

def cell_real_ctype(row,col,sh):
	for merged in sh.merged_cells:
		if (row >= merged[0] and row < merged[1] and col >= merged[2] and col < merged[3]):
			return sh.cell(merged[0],merged[2]).ctype
	return sh.cell(row,col).ctype
	
def readXLSX(filename):
	workbook = xlrd.open_workbook(filename)
	sheets = workbook.sheet_names()
	objlist = []
	for i in range(0, len(sheets)):
		sheet = workbook.sheet_by_name(sheets[i])
		num_rows = sheet.nrows
		num_cols = sheet.ncols
		o = obj()
		for curr_row in range(num_rows):
			for curr_col in range(num_cols):
				ctype = sheet.cell(curr_row, curr_col).ctype
				if ctype == 1:
					value = sheet.cell(curr_row,curr_col).value.strip()
					if value == 'MBL NO':
						o.mbl_no = sheet.cell(curr_row,curr_col + 1).value
					elif value == 'HB/L NO.':
						o.hbl_no = sheet.cell(curr_row,curr_col + 1).value
					elif value == "Container number":
						start_row = curr_row + 1
						t_ctype = sheet.cell(start_row, curr_col).ctype
						while t_ctype == 1 and sheet.cell(start_row,curr_col).value.strip() != "":
							con = container()
							con.container_number = sheet.cell(start_row,curr_col).value
							con.seal_number = sheet.cell(start_row,curr_col + 1).value
							con.amount = sheet.cell(start_row,curr_col + 2).value
							con.type = sheet.cell(start_row,curr_col + 3).value
							con.gross_weight = sheet.cell(start_row,curr_col + 4).value
							con.volume = sheet.cell(start_row,curr_col + 5).value
							o.containers.append(con)
							start_row = start_row + 1
							t_ctype = sheet.cell(start_row, curr_col).ctype
					elif value.find("Cargo description") != -1:
						start_row = curr_row + 1
						while start_row < num_rows and sheet.cell(start_row, curr_col).ctype == 1 and sheet.cell(start_row,curr_col).value.strip() != "":
							o.cargo_descriptions.append(sheet.cell(start_row,curr_col).value)
							start_row = start_row + 1
					elif value == 'HTS code':
						start_row = curr_row + 1
						while start_row < num_rows and (sheet.cell(start_row, curr_col).ctype == 1 or sheet.cell(start_row, curr_col).ctype == 2) and str(sheet.cell(start_row,curr_col).value).strip() != "":
							tt_type = sheet.cell(start_row,curr_col).ctype
							hts_code = sheet.cell(start_row, curr_col).value
							ret_hts_code = ''
							if tt_type == 2:
								if hts_code % 1 == 0:
									ret_hts_code = str(int(hts_code))
								else:
									ret_hts_code = str(hts_code)
							else:
								ret_hts_code = hts_code
							ret_hts_code = ret_hts_code.replace('.', '');
							o.hts_codes.append(ret_hts_code)
							start_row = start_row + 1
					elif value == 'Mark':
						start_row = curr_row + 1
						o.mark_merge = False
						while start_row < num_rows and sheet.cell(start_row, curr_col).ctype == 1 and sheet.cell(start_row,curr_col).value.strip() != "":
							o.marks.append(sheet.cell(start_row, curr_col).value)
							start_row = start_row + 1
						if start_row < num_rows and len(o.marks) == 1 and inMerged(start_row, curr_col, sheet):
							o.mark_merge = True

		if o.mbl_no != "":
			objlist.append(o)
	return objlist

def writeXLSX(objlist, path):
	wb = Workbook()
	ws = wb.active
	ws.cell(row=1, column=1).value = 'MBL NO'
	ws.cell(row=1, column=2).value = 'HB/L NO.'
	ws.cell(row=1, column=3).value = 'Container number'
	ws.cell(row=1, column=4).value = 'Seal number'
	ws.cell(row=1, column=5).value = '数量'.decode('utf-8')
	ws.cell(row=1, column=6).value = '包装类型（CTNS)'.decode('utf-8')
	ws.cell(row=1, column=7).value = 'Gross weight(kg)'
	ws.cell(row=1, column=8).value = 'Volume(CBM)'
	ws.cell(row=1, column=9).value = 'Cargo description'
	ws.cell(row=1, column=10).value = 'HTS code'
	ws.cell(row=1, column=11).value = 'Mark'
	cur_row = 2
	fontObj = Font(name='Maersk Text', bold=False, italic=False, size=10)
	for i in range(len(objlist)):
		ws.cell(row=cur_row, column=1).value = objlist[i].mbl_no
		ws.cell(row=cur_row, column=2).value = objlist[i].hbl_no
		
		ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal='left')
		ws.cell(row=cur_row, column=2).alignment = Alignment(horizontal='left')
		
		ws.cell(row=cur_row, column=1).font = fontObj
		ws.cell(row=cur_row, column=2).font = fontObj
		for m in range(len(objlist[i].containers)):
			ws.cell(row=cur_row + m, column=3).value = objlist[i].containers[m].container_number
			ws.cell(row=cur_row + m, column=4).value = objlist[i].containers[m].seal_number
			ws.cell(row=cur_row + m, column=5).value = objlist[i].containers[m].amount
			ws.cell(row=cur_row + m, column=6).value = objlist[i].containers[m].type
			ws.cell(row=cur_row + m, column=7).value = objlist[i].containers[m].gross_weight
			ws.cell(row=cur_row + m, column=8).value = objlist[i].containers[m].volume
			
			ws.cell(row=cur_row + m, column=3).alignment = Alignment(horizontal='left')
			ws.cell(row=cur_row + m, column=4).alignment = Alignment(horizontal='left')
			ws.cell(row=cur_row + m, column=5).alignment = Alignment(horizontal='left')
			ws.cell(row=cur_row + m, column=6).alignment = Alignment(horizontal='left')
			ws.cell(row=cur_row + m, column=7).alignment = Alignment(horizontal='left')
			ws.cell(row=cur_row + m, column=8).alignment = Alignment(horizontal='left')
			
			ws.cell(row=cur_row + m, column=3).font = fontObj
			ws.cell(row=cur_row + m, column=4).font = fontObj
			ws.cell(row=cur_row + m, column=5).font = fontObj
			ws.cell(row=cur_row + m, column=6).font = fontObj
			ws.cell(row=cur_row + m, column=7).font = fontObj
			ws.cell(row=cur_row + m, column=8).font = fontObj
		for m in range(len(objlist[i].cargo_descriptions)):
			ws.cell(row=cur_row + m, column=9).value = objlist[i].cargo_descriptions[m]
			ws.cell(row=cur_row + m, column=9).alignment = Alignment(horizontal='left')
			ws.cell(row=cur_row + m, column=9).font = fontObj
		for m in range(len(objlist[i].hts_codes)):
			ws.cell(row=cur_row + m, column=10).value = objlist[i].hts_codes[m]
			ws.cell(row=cur_row + m, column=10).alignment = Alignment(horizontal='left')
			ws.cell(row=cur_row + m, column=10).font = fontObj
		for m in range(len(objlist[i].marks)):
			ws.cell(row=cur_row + m, column=11).value = objlist[i].marks[m]
			ws.cell(row=cur_row + m, column=11).alignment = Alignment(horizontal='left')
			ws.cell(row=cur_row + m, column=11).font = fontObj
		
		if objlist[i].mark_merge == True:
			ws.merge_cells(start_row=cur_row, start_column=11, end_row=cur_row+len(objlist[i].hts_codes)-1, end_column=11)
			ws.cell(row=cur_row, column=11).alignment = Alignment(horizontal='left', vertical='center')			
			
		cur_row  = cur_row + max(len(objlist[i].containers), len(objlist[i].cargo_descriptions), len(objlist[i].hts_codes), len(objlist[i].marks))
	autoWidth(ws)
	wb.save(filename=path)
	
if ( __name__ == "__main__"):
	parse_dir = './'
	if os.path.exists('config.ini'):
		remove_BOM("config.ini")
		try:
			cf = ConfigParser.ConfigParser()
			cf.read("config.ini")
			dir = cf.get("config", "parse_dir").strip()
			if dir != "":
				parse_dir = dir
		except:
			print "找不到parse_dir在config.ini文件中".decode("utf-8")
	
	file_removed = True
	output_dir = os.path.join(parse_dir, 'output')
	if not os.path.exists(output_dir):
		os.mkdir(output_dir)
	output_file = os.path.join(output_dir, 'result.xlsx')
	try:
		if os.path.exists(output_file):
			os.remove(output_file)
	except:
		file_removed = False
		print "请关闭result.xlsx文件后重试！".decode('utf-8')
		
	if file_removed:	
		# 遍历并导出所有的Excel文件
		pattern = re.compile('[^~].*?\.(xlsx)$')
		files = os.listdir(parse_dir)
		objlist = []
		total = 0
		cur_num = 0
		faillist = []
		mfiles = []
		for file in files:
			if pattern.match(file):
				total = total + 1;
				mfiles.append(file)
		for file in mfiles:
			cur_num = cur_num + 1
			print '[' + str(cur_num) +  '/' + str(total) + ']Parse File:' + file
			try:
				l = readXLSX(os.path.join(parse_dir, file))
				if len(l) > 0:
					objlist.extend(l)
			except:
				faillist.append(file);
				
		print ('成功: ' + str(total - len(faillist)) + ", 失败: " + str(len(faillist))).decode('utf-8')
		if len(faillist) > 0:
			print '\n[Sorry]部分文件解析失败，请检查文件名是否包含特殊字符、对应表格格式是否正确，或者手动操作这些文件'.decode('utf-8')
			print '###########解析失败文件清单###################'.decode('utf-8')
		for i in range(len(faillist)):
			print faillist[i]
		if len(faillist) > 0:
			print '##########################################'
		writeXLSX(objlist, output_file)
		'''
		for i in range(len(objlist)):
			print objlist[i].__dict__
			for m in range(len(objlist[i].containers)):
				print objlist[i].containers[m].__dict__
				print objlist[i].containers[m].__dict__
			for m in range(len(objlist[i].cargos)):
				print objlist[i].cargos[m].__dict__
		'''
		
		print "\n解析结束，按任意键关闭窗口".decode('utf-8')
	raw_input()
